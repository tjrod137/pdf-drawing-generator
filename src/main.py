import os
import re
from pathlib import Path

import openpyxl

# Acrobat COM (requires: pip install pywin32, and Adobe Acrobat installed)
import win32com.client


INVALID_FILENAME_CHARS = r'[\/:*?"<>|]'


def safe_filename(s: str) -> str:
    """Remove invalid filename characters and trim whitespace."""
    s = "" if s is None else str(s)
    s = s.strip()
    return re.sub(INVALID_FILENAME_CHARS, "", s)


def clean_text(s) -> str:
    """Strip + string conversion + lower not applied here (we preserve original casing for PDFs)."""
    return "" if s is None else str(s).strip()


def is_blank(s) -> bool:
    return s is None or str(s).strip() == ""


def set_field(js_obj, field_name: str, value):
    """Safely set a PDF field value (skip if field missing)."""
    try:
        f = js_obj.getField(field_name)
        if f is not None:
            f.Value = value
    except Exception:
        # In production you might log missing fields; for portfolio, keep it simple.
        pass


def main():
    # --- Paths (edit these for your laptop) ---
    excel_path = r"H:\Documents\Ref_File\Drawing_Entry.xlsx"
    pdf_input = r"H:\Documents\Drawings\Blank_Drawing.pdf"
    save_folder = r"H:\Documents\Drawings\FilledForms"

    os.makedirs(save_folder, exist_ok=True)

    # --- Load Excel (Sheet1 like your script) ---
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb["Sheet1"]  # change if needed

    # Start row (assuming headers in row 1)
    start_row = 2

    for row in range(start_row, sheet.max_row + 1):
        # --- Read values (columns match your AHK script) ---
        name = sheet.cell(row=row, column=1).value
        jobname = sheet.cell(row=row, column=2).value

        address = sheet.cell(row=row, column=3).value
        city = sheet.cell(row=row, column=4).value
        state = sheet.cell(row=row, column=5).value
        zip_code = sheet.cell(row=row, column=6).value

        drawing_name = sheet.cell(row=row, column=7).value

        basin_size = sheet.cell(row=row, column=8).value
        overall_height = sheet.cell(row=row, column=9).value

        chart_selection = sheet.cell(row=row, column=10).value
        grate_style = sheet.cell(row=row, column=11).value

        outlet1 = sheet.cell(row=row, column=12).value
        ot1 = sheet.cell(row=row, column=13).value
        degrees1 = sheet.cell(row=row, column=14).value
        ih1 = sheet.cell(row=row, column=15).value

        outlet2 = sheet.cell(row=row, column=16).value
        ot2 = sheet.cell(row=row, column=17).value
        degrees2 = sheet.cell(row=row, column=18).value
        ih2 = sheet.cell(row=row, column=19).value

        outlet3 = sheet.cell(row=row, column=20).value
        ot3 = sheet.cell(row=row, column=21).value
        degrees3 = sheet.cell(row=row, column=22).value
        ih3 = sheet.cell(row=row, column=23).value

        outlet4 = sheet.cell(row=row, column=24).value
        ot4 = sheet.cell(row=row, column=25).value
        degrees4 = sheet.cell(row=row, column=26).value
        ih4 = sheet.cell(row=row, column=27).value

        # Stop condition (same idea as your AHK)
        if is_blank(name) or is_blank(drawing_name):
            break

        # --- Clean fields / defaults ---
        zip_str = clean_text(zip_code)
        if zip_str == "":
            zip_str = "N/A"

        def clean_degrees(d):
            d = clean_text(d)
            return "" if d == "" else d

        degrees1 = clean_degrees(degrees1)
        degrees2 = clean_degrees(degrees2)
        degrees3 = clean_degrees(degrees3)
        degrees4 = clean_degrees(degrees4)

        # Keep original casing for human-facing PDF fields
        name_str = clean_text(name)
        job_str = clean_text(jobname)
        drawing_str = clean_text(drawing_name)

        # --- Output filename: "Name - Job - Drawing.pdf" ---
        cleaned_name = safe_filename(name_str)
        cleaned_job = safe_filename(job_str)
        cleaned_drawing = safe_filename(drawing_str)

        pdf_output = os.path.join(
            save_folder,
            f"{cleaned_name} - {cleaned_job} - {cleaned_drawing}.pdf"
        )

        # --- Acrobat COM: open PDF and fill ---
        acro_app = win32com.client.Dispatch("AcroExch.App")
        pd_doc = win32com.client.Dispatch("AcroExch.PDDoc")

        opened = pd_doc.Open(pdf_input)
        if not opened:
            acro_app.Exit()
            raise RuntimeError(f"Failed to open PDF template: {pdf_input}")

        js_obj = pd_doc.GetJSObject()

        # Fill fields (names match your AHK)
        set_field(js_obj, "Job Name", job_str)
        set_field(js_obj, "Name", name_str)

        set_field(js_obj, "Address", clean_text(address))
        set_field(js_obj, "City", clean_text(city))
        set_field(js_obj, "State", clean_text(state))
        set_field(js_obj, "Zip", zip_str)

        set_field(js_obj, "Drain Basin / Drawing #", drawing_str)
        set_field(js_obj, "Drain Basin Size*", clean_text(basin_size))
        set_field(js_obj, "Overall Height (A)*", clean_text(overall_height))

        set_field(js_obj, "Outlet Diameter1", clean_text(outlet1))
        set_field(js_obj, "Outlet Type1", clean_text(ot1))
        set_field(js_obj, "Location Degrees1", degrees1)
        set_field(js_obj, "Invert Height B1", clean_text(ih1))

        set_field(js_obj, "Outlet Diameter2", clean_text(outlet2))
        set_field(js_obj, "Outlet Type2", clean_text(ot2))
        set_field(js_obj, "Location Degrees2", degrees2)
        set_field(js_obj, "Invert Height B2", clean_text(ih2))

        set_field(js_obj, "Outlet Diameter3", clean_text(outlet3))
        set_field(js_obj, "Outlet Type3", clean_text(ot3))
        set_field(js_obj, "Location Degrees3", degrees3)
        set_field(js_obj, "Invert Height B3", clean_text(ih3))

        set_field(js_obj, "Outlet Diameter4", clean_text(outlet4))
        set_field(js_obj, "Outlet Type4", clean_text(ot4))
        set_field(js_obj, "Location Degrees4", degrees4)
        set_field(js_obj, "Invert Height B4", clean_text(ih4))

        # Chart selection -> mark Y
        chart_selection = clean_text(chart_selection)
        chart_map = {
            "Round w Frame": "Round w Frame",
            "Drop In": "Drop In",
            "Domed": "Domed",
            "Domed w Frame": "Domed w Frame",
            "Square Hinged 12  15": "Square Hinged 12  15",
            "Curb Inlet 2x2 or 2x3": "Curb Inlet 2x2 or 2x3",
            "Traffic Inlet 2x2 or 2x3": "Traffic Inlet 2x2 or 2x3",
            "No Grate": "No Grate",
        }
        if chart_selection in chart_map:
            set_field(js_obj, chart_map[chart_selection], "Y")

        # Grate style -> mark Y
        grate_style = clean_text(grate_style)
        grate_map = {
            "H10": "H10 Light duty",
            "H25": "H25",
            "Solid": "Solid",
        }
        if grate_style in grate_map:
            set_field(js_obj, grate_map[grate_style], "Y")

        # Calculate + flatten + save (like your AHK)
        try:
            js_obj.calculate
        except Exception:
            pass

        try:
            js_obj.flattenPages()
        except Exception:
            pass

        # Save and close
        pd_doc.Save(1, pdf_output)  # 1 = full save
        pd_doc.Close()

        acro_app.Exit()

        print(f"Saved: {pdf_output}")

    wb.close()
    print("Done: all PDFs created.")


if __name__ == "__main__":
    main()
